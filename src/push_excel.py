import openpyxl
from openpyxl.styles import Font

from config import vacancies_excel
from src.vacancy import Vacancy


class PushExcel:
    """ Класс взаимодействия с Excel документом. """

    def __init__(self, list_vacancies: list[Vacancy]):
        self.list_vacancies = list_vacancies

    def push_excel(self):
        """ Создает Excel-документ и записывает в него вакансии. """
        book = openpyxl.Workbook()
        book.remove(book.active)
        page = book.create_sheet('Вакансии')

        headers = ['ID', 'Вакансия', 'URL', 'Зарплата', 'Дата публикации',
                   'Адрес', 'Компания', 'График работы', 'Тип занятости',
                   'Опыт работы', 'Требования']

        for count, el in enumerate(headers):
            item = page.cell(row=1, column=count + 1, value=el)
            item.font = Font(bold=True)

        row = 2
        column = 1

        for item in self.list_vacancies:
            page.cell(row=row, column=column, value=item.id_)
            page.cell(row=row, column=column + 1, value=item.vacancy)
            page.cell(row=row, column=column + 2, value=item.url)
            page.cell(row=row, column=column + 3, value=item.pay_str)
            page.cell(row=row, column=column + 4, value=item.published)
            page.cell(row=row, column=column + 5, value=item.address)
            page.cell(row=row, column=column + 6, value=item.company)
            page.cell(row=row, column=column + 7, value=item.schedule)
            page.cell(row=row, column=column + 8, value=item.employment)
            page.cell(row=row, column=column + 9, value=item.experience)
            page.cell(row=row, column=column + 10, value=item.snippet)
            row += 1

        book.save(vacancies_excel)
        print('Данные успешно сформированы в Excel-документ в папке "data"!')

    def add_excel(self):
        """ Добавляет вакансии. """
        book = openpyxl.load_workbook(vacancies_excel)
        page = book.active

        row = page.max_row + 1
        column = 1

        for item in self.list_vacancies:
            page.cell(row=row, column=column, value=item.id_)
            page.cell(row=row, column=column + 1, value=item.vacancy)
            page.cell(row=row, column=column + 2, value=item.url)
            page.cell(row=row, column=column + 3, value=item.pay_str)
            page.cell(row=row, column=column + 4, value=item.published)
            page.cell(row=row, column=column + 5, value=item.address)
            page.cell(row=row, column=column + 6, value=item.company)
            page.cell(row=row, column=column + 7, value=item.schedule)
            page.cell(row=row, column=column + 8, value=item.employment)
            page.cell(row=row, column=column + 9, value=item.experience)
            page.cell(row=row, column=column + 10, value=item.snippet)
            row += 1

        book.save(vacancies_excel)

    @staticmethod
    def del_excel(col_id):
        """ Удаляет строку по id. """
        book = openpyxl.load_workbook(filename=vacancies_excel)
        sheet = book.active

        for num_, row in enumerate(sheet.rows):
            if row[0].value == col_id:
                sheet.delete_rows(num_+1)

        book.save(vacancies_excel)
